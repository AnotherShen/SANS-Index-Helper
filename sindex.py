#!/usr/bin/python
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
import sys
import getopt


# Main Function
def main(argv):
    # Init variable
    uniq = {}

    # Command line variables
    infile = ''
    outfile = 'index_out.xlsx'
    try:
        opts, args = getopt.getopt(argv, 'hi:o:', ['ifile=', 'ofile='])
    except getopt.GetoptError:
        print('sindex.py -i <inputfile> [-o <outputfile>]')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('sindex.py -i <inputfile> [-o <outputfile>]')
            sys.exit()
        elif opt in ('-i', '--ifile'):
            if len(arg) == 2 and arg[:2] == '-o':
                print('sindex.py -i <inputfile> [-o <outputfile>]')
                sys.exit(2)
            infile = arg
        elif opt in ('-o', '--ofile'):
            outfile = arg

    # Check that infile has an arg
    if infile == '':
        print('sindex.py -i <inputfile> [-o <outputfile>]')
        sys.exit(2)

    # Add .xlsx if not already in strings
    if not infile[-5:] == '.xlsx':
        infile = infile + '.xlsx'
    if not outfile[-5:] == '.xlsx':
        outfile = outfile + '.xlsx'

    # If input file doesn't exits, close
    if not Path(infile).is_file():
        print(infile+' does not exist!')
        sys.exit(2)

    # Load workbook
    wb = load_workbook(infile)
    ws = wb['Terms']

    # Store all values and pages in a dictionary
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=100, values_only=True):
        # Don't include without a reference
        if str(row[0]) == 'None' or str(row[1]) == 'None':
            break

        # Combine book and page for reference
        reference = str(row[0])+":"+str(row[1])

        # Iterate through terms
        for n in range(2, 5):
            # If term doesn't exist don't include
            if str(row[n]) == 'None':
                break

            # If the term isn't in the dictionary add it, else append the reference
            if row[n] not in uniq:
                uniq[str(row[n])] = reference
            else:
                uniq[str(row[n])] = uniq[str(row[n])] + ", " + reference

    # Close workbook
    wb.close()

    # Write to workbook alphabetically
    wb = Workbook()
    ws = wb.active
    ws.title = 'SANs Index'

    # Add headers
    ws['A1'] = 'Term'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 'Reference'
    ws['B1'].font = Font(bold=True)

    # Populate cells
    curCell = 2
    for key in sorted(uniq):
        ws['A'+str(curCell)] = key
        ws['B'+str(curCell)] = uniq[key]
        curCell += 1

    # Save and close
    wb.save(outfile)
    wb.close()

    # Print success
    print("Success! -> "+infile+" processed into "+outfile)


# Run main
if __name__ == "__main__":
    main(sys.argv[1:])
